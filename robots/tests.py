import json
from datetime import timedelta
from io import BytesIO

from django.http import JsonResponse
from django.test import RequestFactory
from django.test import TestCase
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook

from robots.models import Robot
from robots.utils import get_dataframe
from robots.utils import get_info_robots
from robots.views import ExportToExcel


class TestAddRobot(TestCase):
    @classmethod
    def setUpClass(cls):
        cls.test_data = {
            1: {"model": "R2", "version": "D2", "created": "2022-12-31 23:59:59"},
            2: {"model": "13", "version": "XS", "created": "2023-01-01 00:00:00"},
            3: {"model": "X5", "version": "LT", "created": "2023-01-01 00:00:01"},
        }
        cls.test_invalid_data = {
            1: {"model": 2, "version": "D2", "created": "2022-12-31 23:59:59"},
            2: {"model": "13", "version": "XS", "created": "2023-01-01 00:00:00"},
            3: {"model": "X5", "version": "LT", "created": "2023-01-01 00:00:01"},
        }
        super().setUpClass()

    def test_create_robot(self):
        json_data = json.dumps(self.test_data)
        response: JsonResponse = self.client.post(path=reverse("robots:add"), content_type="json", data=json_data)
        self.assertEqual(response.status_code, 200)

    def test_error_create(self):
        json_data = json.dumps(self.test_invalid_data)
        response: JsonResponse = self.client.post(path=reverse("robots:add"), content_type="json", data=json_data)
        self.assertEqual(response.status_code, 403)


class ExportToExcelTests(TestCase):
    def setUp(self):
        self.factory = RequestFactory()
        self.url = "/export-excel/"

        # Создаем тестовые данные
        now = timezone.now()
        Robot.objects.create(serial="001", model="R2", version="D2", created=now - timedelta(days=3))
        Robot.objects.create(serial="002", model="R2", version="D2", created=now - timedelta(days=5))
        Robot.objects.create(serial="003", model="C3", version="PO", created=now - timedelta(days=6))

    def test_response_metadata(self):
        request = self.factory.get(self.url)
        response = ExportToExcel.as_view()(request)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        self.assertIn("attachment; filename=robots_by_models.xlsx", response["Content-Disposition"])

    def test_excel_structure(self):
        request = self.factory.get(self.url)
        response = ExportToExcel.as_view()(request)

        wb = load_workbook(BytesIO(response.content))

        # Проверяем листы
        self.assertSetEqual(set(wb.sheetnames), {"R2", "C3"})

        # Проверяем заголовки
        r2_sheet = wb["R2"]
        self.assertEqual([cell.value for cell in r2_sheet[1]], ["Модель", "Версия", "Количество за неделю"])

        # Проверяем данные
        data_rows = list(r2_sheet.iter_rows(min_row=2))
        self.assertEqual(len(data_rows), 1)
        self.assertEqual([cell.value for cell in data_rows[0]], ["R2", "D2", 2])


class GetInfoRobotsTests(TestCase):
    def setUp(self):
        # Создаем данные с разными датами
        now = timezone.now()
        Robot.objects.create(model="R2", version="D2", created=now - timedelta(days=6))
        Robot.objects.create(model="R2", version="D2", created=now - timedelta(days=8))  # Старая запись

    def test_returns_only_last_week_data(self):
        result = get_info_robots()
        self.assertEqual(len(result), 1)
        self.assertEqual(result[0].model, "R2")
        self.assertEqual(result[0].version, "D2")
        self.assertEqual(result[0].total, 2)

    def test_aggregation_logic(self):
        # Добавляем еще роботов
        Robot.objects.create(model="R2", version="D2", created=timezone.now())
        Robot.objects.create(model="R2", version="D3", created=timezone.now())

        result = get_info_robots()
        self.assertEqual(len(result), 2)

        versions = {item.version for item in result}
        self.assertSetEqual(versions, {"D2", "D3"})


class GetDataframeTests(TestCase):
    def test_dataframe_creation(self):
        from robots.schemas import RobotToExcel

        test_data = [RobotToExcel(model="R2", version="D2", total=5), RobotToExcel(model="C3", version="PO", total=3)]

        df = get_dataframe(test_data)

        # Проверяем колонки
        self.assertListEqual(list(df.columns), ["Модель", "Версия", "Количество за неделю"])

        # Проверяем данные
        self.assertEqual(df.shape[0], 2)
        self.assertEqual(df.iloc[0]["Модель"], "R2")
        self.assertEqual(df.iloc[1]["Количество за неделю"], 3)
